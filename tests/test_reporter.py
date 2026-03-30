"""Tests pour smb_bleedings.agents.reporter — JSON, CSV et XLSX generation."""

import csv
import io
import json
from datetime import datetime, timezone
from unittest.mock import patch

import pytest

from smb_bleedings.models import AclEntry, ContentMatch, Finding, Host, ScanSummary, Share
from smb_bleedings.agents.reporter import build_summary, _generate_json, _generate_csv


def _make_finding(
    ip="10.0.0.1",
    share_name="Data",
    risk_level="HIGH",
    risk_score=75,
    account="Everyone",
    access_right="Full",
) -> Finding:
    host = Host(ip=ip, hostname="srv01", smb_version="SMB3.1.1", signing_required=True)
    share = Share(host=host, name=share_name, path=f"\\\\{ip}\\{share_name}", tested_permissions=["READ", "LIST"])
    return Finding(
        share=share,
        acl_entries=[AclEntry(account=account, access_right=access_right, ace_type="Allow")],
        dangerous_entries=[AclEntry(account=account, access_right=access_right, ace_type="Allow")],
        risk_level=risk_level,
        risk_score=risk_score,
        reasons=[f"[{share_name}] {account} has {access_right} access"],
        impacts=[f"Anyone can access {share_name}"],
        recommendations=["Restrict access"],
        anonymous_access=False,
    )


# ── build_summary ──

def test_build_summary_counts():
    findings = [
        _make_finding(risk_level="CRITICAL", risk_score=95),
        _make_finding(risk_level="HIGH", risk_score=75),
        _make_finding(risk_level="HIGH", risk_score=70),
        _make_finding(risk_level="MEDIUM", risk_score=40),
    ]
    now = datetime.now(tz=timezone.utc)
    summary = build_summary(findings, ["10.0.0.0/24"], 256, 5, 10, now, now, title="Test", lang="en")
    assert summary.findings_critical == 1
    assert summary.findings_high == 2
    assert summary.findings_medium == 1
    assert summary.title == "Test"
    assert summary.lang == "en"


def test_build_summary_empty_findings():
    now = datetime.now(tz=timezone.utc)
    summary = build_summary([], [], 0, 0, 0, now)
    assert summary.findings_critical == 0
    assert summary.findings_high == 0


# ── _generate_json ──

@patch("smb_bleedings.agents.reporter.pkg_version", return_value="1.2.3")
def test_generate_json_schema(mock_ver, tmp_path):
    findings = [_make_finding()]
    now = datetime.now(tz=timezone.utc)
    summary = build_summary(findings, ["10.0.0.0/24"], 256, 1, 1, now, now, lang="en")
    path = str(tmp_path / "report.json")
    _generate_json(findings, summary, path)

    data = json.loads((tmp_path / "report.json").read_text("utf-8"))
    assert data["meta"]["tool"] == "ShareMyBleedings"
    assert data["meta"]["version"] == "1.2.3"
    assert "generated_at" in data["meta"]
    assert "findings" in data
    assert isinstance(data["findings"], list)
    assert len(data["findings"]) == 1
    assert data["summary"]["high"] >= 1


@patch("smb_bleedings.agents.reporter.pkg_version", return_value="1.0.0")
def test_generate_json_lang_propagated(mock_ver, tmp_path):
    findings = [_make_finding()]
    now = datetime.now(tz=timezone.utc)
    summary = build_summary(findings, [], 0, 0, 0, now, lang="fr")
    path = str(tmp_path / "report.json")
    _generate_json(findings, summary, path)

    data = json.loads((tmp_path / "report.json").read_text("utf-8"))
    assert data["meta"]["lang"] == "fr"


# ── _generate_csv ──

def test_generate_csv_en_uses_comma(tmp_path):
    findings = [_make_finding()]
    path = str(tmp_path / "report.csv")
    _generate_csv(findings, path, lang="en")

    raw = (tmp_path / "report.csv").read_bytes()
    assert raw[:3] == b"\xef\xbb\xbf"  # BOM
    text = raw[3:].decode("utf-8")
    reader = csv.reader(io.StringIO(text), delimiter=",")
    headers = next(reader)
    assert "Risk" in headers
    assert "Score" in headers
    assert "Account" in headers


def test_generate_csv_fr_uses_semicolon(tmp_path):
    findings = [_make_finding()]
    path = str(tmp_path / "report.csv")
    _generate_csv(findings, path, lang="fr")

    raw = (tmp_path / "report.csv").read_bytes()
    text = raw[3:].decode("utf-8")
    reader = csv.reader(io.StringIO(text), delimiter=";")
    headers = next(reader)
    assert "Risque" in headers
    assert "Compte" in headers


def test_generate_csv_content(tmp_path):
    finding = _make_finding(risk_level="CRITICAL", risk_score=95, account="Everyone", access_right="Full")
    path = str(tmp_path / "report.csv")
    _generate_csv([finding], path, lang="en")

    raw = (tmp_path / "report.csv").read_bytes()
    text = raw[3:].decode("utf-8")
    reader = csv.reader(io.StringIO(text), delimiter=",")
    headers = next(reader)
    row = next(reader)
    assert row[0] == "CRITICAL"
    assert row[1] == "95"
    assert "Everyone" in row[8]  # Account column


def test_generate_csv_signing_yes_no(tmp_path):
    finding = _make_finding()
    finding.share.host.signing_required = False
    path = str(tmp_path / "report.csv")
    _generate_csv([finding], path, lang="en")

    raw = (tmp_path / "report.csv").read_bytes()
    text = raw[3:].decode("utf-8")
    assert "No" in text

    finding.share.host.signing_required = True
    _generate_csv([finding], path, lang="fr")
    raw = (tmp_path / "report.csv").read_bytes()
    text = raw[3:].decode("utf-8")
    assert "Oui" in text


def test_generate_csv_empty_findings(tmp_path):
    path = str(tmp_path / "report.csv")
    _generate_csv([], path, lang="en")

    raw = (tmp_path / "report.csv").read_bytes()
    text = raw[3:].decode("utf-8")
    lines = text.strip().split("\n")
    assert len(lines) == 1  # headers only


# ── _generate_xlsx ──

_has_openpyxl = pytest.importorskip is not None
try:
    import openpyxl as _openpyxl
    _has_openpyxl = True
except ImportError:
    _has_openpyxl = False

_skip_xlsx = pytest.mark.skipif(not _has_openpyxl, reason="openpyxl not installed")


@_skip_xlsx
def test_generate_xlsx_creates_file(tmp_path):
    from smb_bleedings.agents.reporter import _generate_xlsx

    findings = [_make_finding()]
    now = datetime.now(tz=timezone.utc)
    summary = build_summary(findings, ["10.0.0.0/24"], 256, 1, 1, now, now, lang="en")
    path = str(tmp_path / "report.xlsx")
    _generate_xlsx(findings, summary, path)
    assert (tmp_path / "report.xlsx").exists()
    assert (tmp_path / "report.xlsx").stat().st_size > 0


@_skip_xlsx
def test_generate_xlsx_has_three_sheets(tmp_path):
    from smb_bleedings.agents.reporter import _generate_xlsx
    import openpyxl

    findings = [_make_finding()]
    now = datetime.now(tz=timezone.utc)
    summary = build_summary(findings, ["10.0.0.0/24"], 256, 1, 1, now, now)
    path = str(tmp_path / "report.xlsx")
    _generate_xlsx(findings, summary, path)

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Summary", "Findings", "Content"]
    wb.close()


@_skip_xlsx
def test_generate_xlsx_findings_rows(tmp_path):
    from smb_bleedings.agents.reporter import _generate_xlsx
    import openpyxl

    findings = [
        _make_finding(risk_level="CRITICAL", risk_score=95),
        _make_finding(ip="10.0.0.2", share_name="Public", risk_level="HIGH", risk_score=65),
    ]
    now = datetime.now(tz=timezone.utc)
    summary = build_summary(findings, [], 0, 0, 2, now, now)
    path = str(tmp_path / "report.xlsx")
    _generate_xlsx(findings, summary, path)

    wb = openpyxl.load_workbook(path)
    ws = wb["Findings"]
    # Header row + 2 findings (1 dangerous entry each)
    assert ws.max_row >= 3
    # First data row should be CRITICAL (sorted by score desc)
    assert ws.cell(row=2, column=1).value == "CRITICAL"
    wb.close()


@_skip_xlsx
def test_generate_xlsx_empty_findings(tmp_path):
    from smb_bleedings.agents.reporter import _generate_xlsx
    import openpyxl

    now = datetime.now(tz=timezone.utc)
    summary = build_summary([], [], 0, 0, 0, now, now)
    path = str(tmp_path / "report.xlsx")
    _generate_xlsx([], summary, path)

    wb = openpyxl.load_workbook(path)
    ws = wb["Findings"]
    assert ws.max_row == 1  # header only
    wb.close()
