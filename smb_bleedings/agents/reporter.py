"""Agent 4 : Génération des rapports HTML, JSON et CSV."""

from __future__ import annotations

import csv
import io
import json
import logging
import webbrowser
from datetime import datetime, timezone
from pathlib import Path

from importlib.metadata import version as pkg_version

from rich.console import Console
from rich.table import Table

from smb_bleedings.models import Finding, ScanSummary

log = logging.getLogger(__name__)
console = Console()

_TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "utils" / "templates"

# UI labels for the HTML report template, by language.
REPORT_LABELS: dict[str, dict[str, str]] = {
    "en": {
        "report_title_default": "SMB Audit Report",
        "report_subtitle": "SMB shares audit report",
        "hosts": "Hosts",
        "shares": "Shares",
        "critical": "Critical", "high": "High", "medium": "Medium", "info": "Info",
        "by_risk": "By risk", "by_host": "By host",
        "search_ph": "Search...", "export_csv": "Export CSV", "results": "result(s)",
        "col_risk": "Risk", "col_score": "Score", "col_machine": "Machine",
        "col_share": "Share", "col_smb": "SMB", "col_signing": "Signing",
        "col_account": "Account", "col_rights": "Rights", "col_reason": "Reason",
        "col_perms": "Tested perms", "copy_unc": "Copy UNC path",
        "concrete_impact": "Concrete impact", "remediation": "Remediation suggestion",
        "why_risky": "Why it's risky", "accessible_to": "Accessible to",
        "legit_access": "Legitimate access", "other_impacts": "Other impacts",
        "command": "Command", "legitimate": "legitimate",
        "signing_ok": "Signing OK", "signing_off": "Signing OFF",
        "report_generated": "Report generated on",
        "perm_list": "Can list every file and folder of the share (names, sizes, dates)",
        "perm_read": "Can read and download the content of every file on the share",
        "perm_write": "Can create, modify or overwrite existing files on the share",
        "perm_delete": "Can permanently delete files and folders from the share",
        "perm_execute": "Can execute programs or scripts directly from the share",
        "perm_special": "Can read or modify the security permissions (ACL) of the share",
        "change_theme": "Toggle theme",
        "collapse": "Collapse / Expand",
        "all_findings": "All findings",
        "who_access": "Who has access?",
        "who_danger": "Risky access",
        "who_legit": "Legitimate access",
        "who_other": "Other accounts",
        "accounts_count": "account(s)",
        "guide_btn": "Guide",
        "guide_title": "How to read this report",
        "guide_close": "Close",
        "by_content": "Sensitive content",
        "col_file": "File",
        "col_keywords": "Keywords",
        "col_size": "Size",
        "no_content_matches": "No sensitive content found",
        "search_content_ph": "Search file or keyword...",
    },
    "fr": {
        "report_title_default": "Rapport d'audit SMB",
        "report_subtitle": "Rapport d'audit des partages SMB",
        "hosts": "Hôtes",
        "shares": "Partages",
        "critical": "Critique", "high": "Élevé", "medium": "Moyen", "info": "Info",
        "by_risk": "Par risque", "by_host": "Par machine",
        "search_ph": "Rechercher...", "export_csv": "Export CSV", "results": "résultat(s)",
        "col_risk": "Risque", "col_score": "Score", "col_machine": "Machine",
        "col_share": "Partage", "col_smb": "SMB", "col_signing": "Signing",
        "col_account": "Compte", "col_rights": "Droits", "col_reason": "Raison",
        "col_perms": "Droits testés", "copy_unc": "Copier le chemin UNC",
        "concrete_impact": "Impact concret", "remediation": "Suggestion de remédiation",
        "why_risky": "Pourquoi c'est risqué", "accessible_to": "Accessible à",
        "legit_access": "Accès légitime", "other_impacts": "Autres impacts",
        "command": "Commande", "legitimate": "légitime",
        "signing_ok": "Signing OK", "signing_off": "Signing OFF",
        "report_generated": "Rapport généré le",
        "perm_list": "Peut lister tous les fichiers et dossiers du partage (noms, tailles, dates)",
        "perm_read": "Peut lire et télécharger le contenu de tous les fichiers du partage",
        "perm_write": "Peut créer, modifier ou écraser des fichiers existants sur le partage",
        "perm_delete": "Peut supprimer des fichiers et dossiers du partage de manière irréversible",
        "perm_execute": "Peut exécuter des programmes ou scripts directement depuis le partage",
        "perm_special": "Peut lire ou modifier les permissions de sécurité (ACL) du partage",
        "change_theme": "Changer de thème",
        "collapse": "Réduire / Agrandir",
        "all_findings": "Tous les findings",
        "who_access": "Qui a accès ?",
        "who_danger": "Accès risqué",
        "who_legit": "Accès légitime",
        "who_other": "Autres comptes",
        "accounts_count": "compte(s)",
        "guide_btn": "Guide",
        "guide_title": "Comment lire ce rapport",
        "guide_close": "Fermer",
        "by_content": "Contenu sensible",
        "col_file": "Fichier",
        "col_keywords": "Mots-clés",
        "col_size": "Taille",
        "no_content_matches": "Aucun contenu sensible détecté",
        "search_content_ph": "Rechercher fichier ou mot-clé...",
    },
}


def _is_wsl() -> bool:
    """Détecte si on tourne sous WSL."""
    try:
        return "microsoft" in Path("/proc/version").read_text().lower()
    except Exception:
        return False


def _open_report(html_path: str) -> None:
    """Ouvre le rapport HTML dans le navigateur, compatible WSL."""
    import subprocess

    resolved = str(Path(html_path).resolve())

    if _is_wsl():
        # Convertir le chemin Linux en chemin Windows pour explorer.exe
        try:
            win_path = subprocess.check_output(
                ["wslpath", "-w", resolved], text=True
            ).strip()
            # Redirect stdin/stdout/stderr to DEVNULL to prevent
            # explorer.exe from corrupting the WSL terminal (echo off bug)
            subprocess.Popen(
                ["explorer.exe", win_path],
                stdin=subprocess.DEVNULL,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            # Restore terminal sanity — explorer.exe under WSL can
            # mess up terminal settings (disable echo, raw mode, etc.)
            try:
                subprocess.run(["stty", "sane"], check=False)
            except Exception:
                pass
            console.print("\n[cyan]  Rapport ouvert dans le navigateur Windows")
            return
        except Exception as exc:
            log.debug("WSL browser open failed: %s", exc)

    try:
        webbrowser.open(f"file://{resolved}")
        console.print("\n[cyan]  Rapport ouvert dans le navigateur")
    except Exception:
        console.print(f"\n[dim]  Ouvrir manuellement : {resolved}")


def _finding_to_dict(f: Finding) -> dict:
    """Convertit un Finding en dict sérialisable."""
    return {
        "risk_level": f.risk_level,
        "risk_score": f.risk_score,
        "host": {
            "ip": f.share.host.ip,
            "hostname": f.share.host.hostname,
            "smb_version": f.share.host.smb_version,
            "signing_required": f.share.host.signing_required,
        },
        "share": {
            "name": f.share.name,
            "unc_path": f.share.path,
            "description": f.share.description,
            "share_type": f.share.share_type,
            "anonymous_access": f.anonymous_access,
            "tested_permissions": f.share.tested_permissions,
        },
        "acl": [
            {
                "account": e.account,
                "access_right": e.access_right,
                "ace_type": e.ace_type,
            }
            for e in f.dangerous_entries
        ],
        "all_acl": [
            {
                "account": e.account,
                "access_right": e.access_right,
                "ace_type": e.ace_type,
            }
            for e in f.acl_entries
        ],
        "reasons": f.reasons,
        "impacts": f.impacts,
        "recommendations": f.recommendations,
        "content_matches": [
            {
                "file_path": m.file_path,
                "file_size": m.file_size,
                "matched_keywords": m.matched_keywords,
                "local_loot_path": Path(m.local_loot_path).name if m.local_loot_path else None,
                "sha256": m.sha256,
            }
            for m in f.content_matches
        ],
    }


def _generate_html(findings: list[Finding], summary: ScanSummary, path: str) -> None:
    """Génère le rapport HTML via template Jinja2."""
    from jinja2 import Environment, FileSystemLoader

    env = Environment(
        loader=FileSystemLoader(str(_TEMPLATE_DIR)),
        autoescape=True,
    )
    import re as _re
    env.filters["strftime"] = lambda dt, fmt: dt.strftime(fmt) if dt else ""
    env.filters["regex_replace"] = lambda s, pattern, repl="": _re.sub(pattern, repl, s) if s else ""

    lang = (summary.lang or "en").lower()
    if lang not in REPORT_LABELS:
        lang = "en"
    L = REPORT_LABELS[lang]

    template = env.get_template("report.html.j2")
    html = template.render(
        findings=sorted(findings, key=lambda f: f.risk_score, reverse=True),
        summary=summary,
        generated_at=datetime.now(tz=timezone.utc),
        lang=lang,
        L=L,
    )
    Path(path).write_text(html, encoding="utf-8")


def _generate_json(
    findings: list[Finding], summary: ScanSummary, path: str
) -> None:
    """Génère le rapport JSON."""
    data = {
        "meta": {
            "tool": "ShareMyBleedings",
            "version": pkg_version("sharemybleedings"),
            "title": summary.title or "",
            "generated_at": datetime.now(tz=timezone.utc).isoformat(),
            "duration_seconds": summary.duration_seconds,
            "ranges": summary.ranges_scanned,
            "total_ips_scanned": summary.total_ips,
            "hosts_discovered": summary.hosts_discovered,
            "lang": (summary.lang or "en").lower(),
        },
        "summary": {
            "critical": summary.findings_critical,
            "high": summary.findings_high,
            "medium": summary.findings_medium,
            "info": summary.findings_info,
            "total_findings": len(findings),
            "hosts_with_findings": summary.hosts_with_findings,
        },
        "findings": [_finding_to_dict(f) for f in findings],
    }
    Path(path).write_text(
        json.dumps(data, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )


def _generate_csv(findings: list[Finding], path: str, lang: str = "en") -> None:
    """Génère le rapport CSV (UTF-8 BOM). Séparateur `;` en FR, `,` sinon."""
    delim = ";" if lang == "fr" else ","
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delim, quoting=csv.QUOTE_MINIMAL)
    if lang == "fr":
        writer.writerow([
            "Risque", "Score", "Machine", "IP", "SMB Version", "Signing",
            "Partage", "UNC", "Compte", "Droits", "Type ACE",
            "Accès anonyme", "Droits testes", "Raison",
            "Contenu sensible (nb)", "Keywords matches", "Fichiers sensibles",
        ])
    else:
        writer.writerow([
            "Risk", "Score", "Machine", "IP", "SMB Version", "Signing",
            "Share", "UNC", "Account", "Rights", "ACE Type",
            "Anonymous access", "Tested permissions", "Reason",
            "Sensitive content (count)", "Matched keywords", "Sensitive files",
        ])

    _yes = "Oui" if lang == "fr" else "Yes"
    _no = "Non" if lang == "fr" else "No"
    for f in sorted(findings, key=lambda x: x.risk_score, reverse=True):
        sign_str = _yes if f.share.host.signing_required else _no if f.share.host.signing_required is False else ""
        cms = f.content_matches or []
        cm_count = len(cms)
        cm_keywords = ""
        cm_files = ""
        if cms:
            kws: set[str] = set()
            for m in cms:
                for k in (m.matched_keywords or []):
                    kws.add(k)
            cm_keywords = "|".join(sorted(kws))
            cm_files = "|".join((m.file_path or "") for m in cms)
        for entry in f.dangerous_entries:
            writer.writerow([
                f.risk_level,
                f.risk_score,
                f.share.host.hostname or "",
                f.share.host.ip,
                f.share.host.smb_version or "",
                sign_str,
                f.share.name,
                f.share.path,
                entry.account,
                entry.access_right,
                entry.ace_type,
                _yes if f.anonymous_access else _no,
                ", ".join(f.share.tested_permissions),
                "; ".join(f.reasons),
                cm_count,
                cm_keywords,
                cm_files,
            ])

    Path(path).write_bytes(b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8"))


def _generate_xlsx(findings: list[Finding], summary: ScanSummary, path: str) -> None:
    """Génère un rapport Excel (XLSX) multi-feuilles avec mise en forme.

    Feuilles :
      - Summary : méta-infos + compteurs par niveau
      - Findings : une ligne par (finding × ACE dangereuse), colonnes enrichies
      - Content : une ligne par fichier sensible détecté (file UNC, hash, keywords…)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        console.print(
            "[yellow]  openpyxl non installé — skip XLSX. "
            "Installer via: pip install 'sharemybleedings[xlsx]' ou pip install openpyxl[/]"
        )
        return

    wb = Workbook()

    # ── Styles communs ──
    HEADER_FILL = PatternFill("solid", fgColor="1F2937")
    HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="F1F5F9")
    HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
    THIN = Side(border_style="thin", color="E5E7EB")
    CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    LEVEL_FILLS = {
        "CRITICAL": PatternFill("solid", fgColor="FCA5A5"),
        "HIGH": PatternFill("solid", fgColor="FDBA74"),
        "MEDIUM": PatternFill("solid", fgColor="FDE68A"),
        "INFO": PatternFill("solid", fgColor="BAE6FD"),
    }
    LEVEL_FONT = {
        "CRITICAL": Font(bold=True, color="7F1D1D"),
        "HIGH": Font(bold=True, color="7C2D12"),
        "MEDIUM": Font(bold=True, color="713F12"),
        "INFO": Font(bold=True, color="0C4A6E"),
    }

    def _style_header(ws, ncols: int) -> None:
        for col in range(1, ncols + 1):
            c = ws.cell(row=1, column=col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = HEADER_ALIGN
            c.border = CELL_BORDER
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    def _autosize(ws, max_widths: list[int]) -> None:
        for i, w in enumerate(max_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet: Summary ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    dur = summary.duration_seconds
    dur_str = f"{int(dur // 60)}m{int(dur % 60):02d}s" if dur >= 60 else f"{dur:.1f}s"
    total_cm = sum(len(f.content_matches or []) for f in findings)

    # Executive summary narrative
    total_findings = len(findings)
    top3 = sorted(findings, key=lambda f: f.risk_score, reverse=True)[:3]
    top3_text = "; ".join(
        f"{f.share.path} ({f.risk_level}, score {f.risk_score})" for f in top3
    ) if top3 else "None"

    exec_summary = (
        f"{summary.hosts_discovered} SMB hosts scanned across {summary.total_ips} IPs. "
        f"{summary.findings_critical} critical and {summary.findings_high} high-risk findings "
        f"require immediate attention. "
        f"Top priorities: {top3_text}."
    )

    meta_rows = [
        ("Executive Summary", exec_summary),
        ("", ""),
        ("Report", summary.title or "SMB Audit Report"),
        ("Generated at", datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Scan started", summary.started_at.strftime("%Y-%m-%d %H:%M:%S") if summary.started_at else ""),
        ("Duration", dur_str),
        ("Ranges", ", ".join(summary.ranges_scanned or [])),
        ("", ""),
        ("IPs scanned", summary.total_ips),
        ("Hosts SMB discovered", summary.hosts_discovered),
        ("Hosts with findings", summary.hosts_with_findings),
        ("Shares analyzed", summary.shares_analyzed),
        ("", ""),
        ("CRITICAL findings", summary.findings_critical),
        ("HIGH findings", summary.findings_high),
        ("MEDIUM findings", summary.findings_medium),
        ("INFO findings", summary.findings_info),
        ("Total findings", total_findings),
        ("", ""),
        ("Sensitive files detected", total_cm),
    ]
    ws_sum["A1"] = "Field"
    ws_sum["B1"] = "Value"
    _style_header(ws_sum, 2)
    for i, (k, v) in enumerate(meta_rows, start=2):
        cell_a = ws_sum.cell(row=i, column=1, value=k)
        cell_b = ws_sum.cell(row=i, column=2, value=v)
        if k == "Executive Summary":
            cell_a.font = Font(bold=True, color="1F2937", size=11)
            cell_b.font = Font(color="1F2937", size=10)
            cell_b.alignment = Alignment(wrap_text=True)
        else:
            cell_a.font = Font(bold=True, color="475569")
    _autosize(ws_sum, [28, 80])

    # Remediation guidance by risk level
    REMEDIATION_TEXT = {
        "CRITICAL": "Remediate IMMEDIATELY — remove dangerous permissions now",
        "HIGH": "Remediate within 1 week — high exposure risk",
        "MEDIUM": "Schedule for next change window — review necessity",
        "INFO": "No action required — informational only",
    }

    # Score conditional formatting fills
    SCORE_FILLS = {
        "critical": PatternFill("solid", fgColor="FCA5A5"),  # 80-100
        "high": PatternFill("solid", fgColor="FDBA74"),      # 60-79
        "medium": PatternFill("solid", fgColor="FDE68A"),    # 40-59
        "low": PatternFill("solid", fgColor="BAE6FD"),       # 0-39
    }

    # ── Sheet: Findings ──
    ws_f = wb.create_sheet("Findings")
    headers = [
        "Risk", "Score", "Machine", "IP", "Share", "UNC Path",
        "Account", "Rights", "ACE Type",
        "Urgency", "Recommendations", "Reasons",
        "SMB Version", "Signing", "Anonymous", "Tested perms",
        "Content matches (nb)", "Matched keywords", "Sensitive files",
    ]
    ws_f.append(headers)
    _style_header(ws_f, len(headers))

    for f in sorted(findings, key=lambda x: x.risk_score, reverse=True):
        sign_str = (
            "Yes" if f.share.host.signing_required
            else "No" if f.share.host.signing_required is False else ""
        )
        cms = f.content_matches or []
        cm_count = len(cms)
        kws: set[str] = set()
        for m in cms:
            for k in m.matched_keywords or []:
                kws.add(k)
        cm_keywords = ", ".join(sorted(kws))
        cm_files = "\n".join(m.file_path or "" for m in cms)
        urgency = REMEDIATION_TEXT.get(f.risk_level, "")
        recs_text = "\n".join(f.recommendations or [])

        entries = f.dangerous_entries or [None]
        for entry in entries:
            row = [
                f.risk_level,
                f.risk_score,
                f.share.host.hostname or "",
                f.share.host.ip,
                f.share.name,
                f.share.path,
                entry.account if entry else "",
                entry.access_right if entry else "",
                entry.ace_type if entry else "",
                urgency,
                recs_text,
                "\n".join(f.reasons or []),
                f.share.host.smb_version or "",
                sign_str,
                "Yes" if f.anonymous_access else "No",
                ", ".join(f.share.tested_permissions),
                cm_count,
                cm_keywords,
                cm_files,
            ]
            ws_f.append(row)
            r = ws_f.max_row
            # Risk level cell styling
            lvl_cell = ws_f.cell(row=r, column=1)
            if f.risk_level in LEVEL_FILLS:
                lvl_cell.fill = LEVEL_FILLS[f.risk_level]
                lvl_cell.font = LEVEL_FONT[f.risk_level]
                lvl_cell.alignment = Alignment(horizontal="center", vertical="center")
            # Score cell conditional coloring
            score_cell = ws_f.cell(row=r, column=2)
            if f.risk_score >= 80:
                score_cell.fill = SCORE_FILLS["critical"]
            elif f.risk_score >= 60:
                score_cell.fill = SCORE_FILLS["high"]
            elif f.risk_score >= 40:
                score_cell.fill = SCORE_FILLS["medium"]
            else:
                score_cell.fill = SCORE_FILLS["low"]
            score_cell.font = Font(bold=True)
            score_cell.alignment = Alignment(horizontal="center", vertical="center")
            # All cells: border + wrap
            for col in range(1, len(headers) + 1):
                cc = ws_f.cell(row=r, column=col)
                cc.alignment = Alignment(vertical="top", wrap_text=True)
                cc.border = CELL_BORDER

    _autosize(ws_f, [11, 7, 20, 15, 22, 40, 28, 14, 10, 30, 50, 50, 12, 9, 11, 22, 10, 30, 50])
    ws_f.auto_filter.ref = ws_f.dimensions
    # Freeze first 2 columns (Risk + Score) for scrolling context
    ws_f.freeze_panes = "C2"

    # ── Sheet: Content matches ──
    ws_c = wb.create_sheet("Content")
    c_headers = [
        "Risk", "Score", "Machine", "IP", "Share",
        "File path", "File UNC", "Size (B)", "Keywords", "SHA256", "Local loot",
    ]
    ws_c.append(c_headers)
    _style_header(ws_c, len(c_headers))

    for f in sorted(findings, key=lambda x: x.risk_score, reverse=True):
        for m in f.content_matches or []:
            file_rel = (m.file_path or "").replace("/", "\\").lstrip("\\")
            file_unc = f"\\\\{f.share.host.ip}\\{f.share.name}\\{file_rel}"
            row = [
                f.risk_level,
                f.risk_score,
                f.share.host.hostname or "",
                f.share.host.ip,
                f.share.name,
                m.file_path or "",
                file_unc,
                m.file_size or 0,
                ", ".join(m.matched_keywords or []),
                m.sha256 or "",
                Path(m.local_loot_path).name if m.local_loot_path else "",
            ]
            ws_c.append(row)
            r = ws_c.max_row
            lvl_cell = ws_c.cell(row=r, column=1)
            if f.risk_level in LEVEL_FILLS:
                lvl_cell.fill = LEVEL_FILLS[f.risk_level]
                lvl_cell.font = LEVEL_FONT[f.risk_level]
                lvl_cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, len(c_headers) + 1):
                cc = ws_c.cell(row=r, column=col)
                cc.alignment = Alignment(vertical="top", wrap_text=True)
                cc.border = CELL_BORDER

    _autosize(ws_c, [11, 7, 20, 15, 22, 55, 60, 12, 40, 50, 55])
    if ws_c.max_row >= 2:
        ws_c.auto_filter.ref = ws_c.dimensions

    wb.save(path)


def _print_summary_table(summary: ScanSummary) -> None:
    """Affiche le tableau récapitulatif Rich."""
    from rich.panel import Panel

    # Format duration
    dur = summary.duration_seconds
    if dur >= 60:
        dur_str = f"{int(dur // 60)}m{int(dur % 60):02d}s"
    else:
        dur_str = f"{dur:.1f}s"

    # Scan stats
    lines: list[str] = [
        f"  [dim]Durée[/]       {dur_str}",
        f"  [dim]IPs[/]         {summary.total_ips}",
        f"  [dim]Hôtes SMB[/]   {summary.hosts_discovered}",
        f"  [dim]Partages[/]    {summary.shares_analyzed}",
    ]

    # Findings breakdown
    total_findings = summary.findings_critical + summary.findings_high + summary.findings_medium + summary.findings_info
    lines.append("")

    if total_findings:
        # Risk bar: visual block representation
        bars = (
            "[bold red]" + "█" * summary.findings_critical + "[/]"
            + "[yellow]" + "█" * summary.findings_high + "[/]"
            + "[blue]" + "█" * summary.findings_medium + "[/]"
            + "[dim]" + "█" * summary.findings_info + "[/]"
            + "[green]" + "░" * min(summary.findings_ok, 20) + "[/]"
        )
        lines.append(f"  {bars}")
        lines.append("")
        if summary.findings_critical:
            lines.append(f"  [bold red]{summary.findings_critical:>3}[/]  [bold red]CRITICAL[/]")
        if summary.findings_high:
            lines.append(f"  [yellow]{summary.findings_high:>3}[/]  HIGH")
        if summary.findings_medium:
            lines.append(f"  [blue]{summary.findings_medium:>3}[/]  MEDIUM")
        if summary.findings_info:
            lines.append(f"  [dim]{summary.findings_info:>3}[/]  INFO")
        if summary.findings_ok:
            lines.append(f"  [green]{summary.findings_ok:>3}[/]  OK")
    else:
        lines.append("  [bold green]Aucun finding — réseau propre[/]")

    panel = Panel(
        "\n".join(lines),
        title="[bold red]SMBleedings[/] [dim]Résultats[/]",
        border_style="red",
        padding=(1, 2),
    )
    console.print(panel)


def build_summary(
    findings: list[Finding],
    ranges: list[str],
    total_ips: int,
    hosts_discovered: int,
    shares_total: int,
    started_at: datetime,
    finished_at: datetime | None = None,
    title: str = "",
    lang: str = "en",
) -> ScanSummary:
    """Construit un ScanSummary à partir des findings."""
    finished = finished_at or datetime.now(tz=timezone.utc)
    host_ips = {f.share.host.ip for f in findings}
    return ScanSummary(
        title=title,
        started_at=started_at,
        finished_at=finished,
        duration_seconds=(finished - started_at).total_seconds(),
        ranges_scanned=ranges,
        total_ips=total_ips,
        hosts_discovered=hosts_discovered,
        hosts_with_findings=len(host_ips),
        shares_total=shares_total,
        shares_analyzed=shares_total,
        findings_critical=sum(1 for f in findings if f.risk_level == "CRITICAL"),
        findings_high=sum(1 for f in findings if f.risk_level == "HIGH"),
        findings_medium=sum(1 for f in findings if f.risk_level == "MEDIUM"),
        findings_info=sum(1 for f in findings if f.risk_level == "INFO"),
        findings_ok=shares_total - len(findings),
        lang=lang,
    )


def generate_report(
    findings: list[Finding],
    summary: ScanSummary,
    output_path: str,
    fmt: str = "html",
    open_browser: bool = True,
) -> dict[str, str]:
    """Génère le(s) rapport(s) demandé(s). Retourne {format: chemin}."""
    base = Path(output_path)
    stem = base.stem
    parent = base.parent
    parent.mkdir(parents=True, exist_ok=True)

    results: dict[str, str] = {}
    formats = ["html", "json", "csv", "xlsx"] if fmt == "all" else [fmt]

    for f in formats:
        path = str(parent / f"{stem}.{f}") if fmt == "all" else str(base)
        if f == "html":
            _generate_html(findings, summary, path)
        elif f == "json":
            _generate_json(findings, summary, path)
        elif f == "csv":
            _generate_csv(findings, path, lang=summary.lang)
        elif f == "xlsx":
            _generate_xlsx(findings, summary, path)
        results[f] = path
        console.print(f"  [green]{f.upper():<4}[/]  {path}")

    _print_summary_table(summary)

    if open_browser and "html" in results:
        _open_report(results["html"])

    return results
